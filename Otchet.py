import os
import pickle
import openpyxl as opx
import datetime
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
import itertools



date_start = datetime.datetime(2000, 1, 1)
date_end = datetime.datetime(2020, 12, 15)
worker = {}
replaced_worker = {}
themes = []
tema_report = {}
unknown_thems = []
svod_row = 0
production_row = {}
report_row = 0
svod_data = []
production_data = {}
prod_names = []
report_data = []


good_names = {'Катя': 'Шаповал Е.С.', 'Женя': 'Гусева Е.Н.', 'Дима': 'Пихуров Д.В.', 'Вова': 'Васильев В.А.',
              'Иван': 'Зуев И.А.', 'Артур': 'Калимуллин А.В.'}
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))



# Меняет даты
def change_date(start, end):
    global date_start, date_end
    date_start = datetime.datetime(start[0], start[1], start[2])
    date_end = datetime.datetime(end[0], end[1], end[2])


def path_to_excel(name: str) -> None:
    """
    Находит путь к файлу эксель
    :return:
    """

    pass


def read_excel(ws, col: int = 5) -> list:
    """
    Функция для чтения excel файла
    :param ws: Лист страницы Excel, с которого надо получить данные
    :param col: количество считаных столбцов
    :return: Список массивов с элементами ячеек
    """
    sv_tabl = []  # Список массивов, который вернем

    for row in ws:  # Проходим по всем строкам на листе
        current_row = []  # Память для текущего ряда

        # Проходим по всем ячейкам в строке. Нумеруем, что бы не считывать все колонки
        for index, cell in enumerate(row):
            if index == col:
                break
            current_row.append(cell.value)  # Добавляем значение ячейки в массив
        if any(current_row):
            sv_tabl.append(current_row)  # Добавляем массив в список
    return sv_tabl


# Форматирует ФИО к стандартному формату
def input_name(name: str) -> str:
    """
    :param name: str (ФамилияИ.О. в любом формате)
    :return: str (Фамилия И.О. с пробелом после фамилии)
    """

    integ = ''  # Возвращаемая строка
    probel_check = False  # Флаг для постановки пробела после фамилии
    prev = ''  # Предыдущий символ при проходе ФИО
    for index, i in enumerate(name):  # Проходим по всем элементам
        if i == '.' and not probel_check:  # Если встречаем точку
            integ += ' '  # Вставляем пробел перед добавлением предыдущего символа
            probel_check = True   # Меняем флаг
        if i != ' ':  # Если символ не пробел
            integ += prev  # Добавляем предыдущий
            prev = i  # Запоминаем текущий
    integ += prev  # Добавляем последний в конце
    return integ  # Возвращаем ФИО


# Обработка данных из сводной таблицы
def svod_tabl_count(tabl: list):
    """
    Элементы массива (номер колонки - 1):
    0) Маркировка плёнки
    1) Тема
    2) Нанесение
    3) Дата
    4) Ответственный
    Добавляет в общий отчет данные из сводной таблицы
    :param tabl: Полученная таблица
    :return: А хз, еще не решил
    """
    global worker, themes

    for row in tabl:  # Берем строку
        if type(row[1]) is not str:
            row[1] = str(row[1])
        if type(row[3]) is type(date_start):  # Смотрим, является ли клетка с датой датой, а не заглавной строкой
            if date_start <= row[3] <= date_end:  # Смотрим, находится ли дата в нужном диапазоне
                if row[4]:
                    row[4] = input_name(row[4])  # Форматируем ФИО по шаблону
                if row[4] not in worker:  # Если нет работника в словаре, то создаем
                    worker[row[4]] = {}
                if row[1] not in worker[row[4]]:  # Если нет темы у работника, то создаем
                    worker[row[4]][row[1]] = {'образцы': [], 'плёнки': [], 'отчёты': []}
                worker[row[4]][row[1]]['плёнки'].append(row[0])  # Добавляем номер пленки работнику в тему
                if row[1] not in themes and row[1]:
                    themes.append(row[1])

    return None


# # Обработка данных из перечня отчётов
def report_count(tabl):
    """
    0) Номер отчёта
    1) Название отчёта
    2) Ответственный 1
    3) Ответственный 2
    4) Ответственный 3
    5) Дата
    6) Название темы
    :param tabl:
    :return:
    """
    for row in tabl:  # Берем строку
        if type(row[6]) is not str:
            row[6] = str(row[6])
        if type(row[5]) is type(date_start):  # Смотрим, является ли клетка с датой датой, а не заглавной строкой
            if date_start <= row[5] <= date_end:  # Смотрим, находится ли дата в нужном диапазоне
                row[2] = input_name(row[2])  # Форматируем ФИО по шаблону
                if row[3]:
                    row[3] = input_name(row[3])
                if row[4]:
                    row[4] = input_name(row[4])

                if row[2] not in worker:  # Если нет работника в словаре, то создаем
                    worker[row[2]] = {}
                if row[6] not in worker[row[2]]:  # Если нет темы у работника, то создаем
                    worker[row[2]][row[6]] = {'образцы': [], 'плёнки': [], 'отчёты': []}
                if row[0]:
                    worker[row[2]][row[6]]['отчёты'].append(row[0])  # Добавляем отчёт
                if row[6] not in themes and row[6]:   # Добавляем тему в список тем, если её там еще нет
                    themes.append(row[6])


# Обработка данных из перечня продукции
def production_count(tabl: list, name: str):
    """
    Элементы массива (номер колонки - 1):
    0) Маркировка состава
    1) Дата
    2) Тема
    :param tabl: Полученная таблица
    :param name: Имя сотрудника
    :return: None
    """
    global worker
    col_flag = [False, False, False]
    mark_col = 0
    date_col = 1
    tema_col = 2

    for row in tabl:  # Берем строку
        for col, cell in enumerate(row):
            if cell and not all(col_flag):
                if type(cell) is str and 'аркиров' in cell:
                    mark_col = col
                    col_flag[0] = True
                if type(cell) is str and ('Дата' in cell or 'дата' in cell):
                    date_col = col
                    col_flag[1] = True
                if type(cell) is str and ('именован' in cell or 'истем' in cell):
                    tema_col = col
                    col_flag[2] = True
        if type(row[tema_col]) is not str and all(col_flag):
            row[tema_col] = str(row[tema_col])
        if type(row[date_col]) is type(date_start):  # Смотрим, является ли клетка с датой датой, а не заглавной строкой
            if date_start <= row[date_col] <= date_end:
                if name not in worker:  # Если нет работника в словаре, то создаем
                    worker[name] = {}
                if row[tema_col]:  # Если клетка с темой не пустая
                    if row[tema_col] not in worker[name]:  # Если нет темы у работника, то создаем
                        worker[name][row[tema_col]] = {'образцы': [], 'плёнки': [], 'отчёты': []}
                    # Добавляем номер пленки работнику в тему
                    worker[name][row[tema_col]]['образцы'].append(row[mark_col])
                if row[tema_col] not in themes and row[tema_col]:
                    themes.append(row[tema_col])
    return None


def reader(svod='Сводная таблица.xlsm',
              production='Общий перечень продукции ОВНТ.xlsx', report='Общий перечень отчётов.xlsm', force_read=False):
    global worker, themes, svod_data, prod_names, production_data, report_data, svod_row, production_row, report_row


    # load_data()
    progressbar = 10
    yield progressbar

    if not svod_data or force_read:
        svod_wb = opx.load_workbook(filename=svod)
        svod_ws = svod_wb.active
        svod_data = read_excel(svod_ws)

    progressbar += 20
    yield progressbar

    if not report_data or force_read:
        report_wb = opx.load_workbook(filename=report)
        report_ws = report_wb.active
        report_data = read_excel(report_ws, col=7)

    progressbar += 20
    yield progressbar

    if not production_data or force_read:
        production_wb = opx.load_workbook(filename=production)
        prod_names = production_wb.sheetnames
        for name in prod_names:
            if name not in worker:
                if name not in good_names:
                    continue
            production_data[name] = read_excel(production_wb[name], col=5)
            progressbar += 5
            yield progressbar

    yield  100
    # save_data()



# Управляет обработчиками
def conductor():
    """
    Функция читает все файлы и прогоняет по ним обсчитыающие функции
    :param svod: Имя файла "Сводная таблица"
    :param production: Имя файла "Перечень продукции ОВНТ"
    :param report: Имя файла "Общий перечень отчётов"
    :return: None
    """

    global worker, themes, tema_report, unknown_thems

    worker = {}
    themes = []
    tema_report = {}
    unknown_thems = []

    svod_tabl_count(svod_data)

    for name in production_data:
        production_count(production_data[name], good_names[name])

    report_count(report_data)



# Сворачивает плёнки и образцы в компактный вид
def short_show(massiv):
    """
    Функция позволяет сворачивать штучные образцы в дефисные группы
    :param massiv: Массив с образцами или плёнками поштучно
    :return: Массив с образцами или плёнками, свёрнуто дефисными группами
    """
    fine = []
    current_prefix = ''
    current_numb = -1
    first_numb = -1
    for word in massiv:
        word_prefix = ''
        word_numb = ''
        for b in word:
            if b.isdigit():
                word_numb += b
            else:
                word_prefix += b
        if word_prefix == current_prefix and int(current_numb) == int(word_numb) - 1:
            current_numb = word_numb
            if word == massiv[-1]:
                if current_numb != first_numb:
                    fine.append(f'{current_prefix}{first_numb}-{current_numb}')
                else:
                    fine.append(f'{current_prefix}{first_numb}')
        else:
            if current_prefix:
                if current_numb != first_numb:
                    fine.append(f'{current_prefix}{first_numb}-{current_numb}')
                else:
                    fine.append(f'{current_prefix}{first_numb}')
            current_prefix = word_prefix
            first_numb = word_numb
            current_numb = first_numb
            if word == massiv[-1] and current_prefix:
                if current_numb != first_numb:
                    fine.append(f'{current_prefix}{first_numb}-{current_numb}')
                else:
                    fine.append(f'{current_prefix}{first_numb}')
    return fine


# Переделывает данные отталкиваясь от темы
def worker_into_thems():
    global tema_report
    for tema in themes:  # Проходим по всем темам
        for name in replaced_worker:  # Проходим по всем сотрудникам
            if tema in replaced_worker[name]:  # Смотрим есть ли тема у сотрудника
                if tema not in tema_report:  # Проверяем есть ли тема в отчётном словаре
                    tema_report[tema] = {}
                if name not in tema_report[tema]:  # Проверяем есть ли сотрудник в этой теме
                    tema_report[tema][name] = {'образцы': [], 'плёнки': [], 'отчёты': []}
                for i in replaced_worker[name][tema]['образцы']:
                    tema_report[tema][name]['образцы'].append(i)
                for i in replaced_worker[name][tema]['плёнки']:
                    tema_report[tema][name]['плёнки'].append(i)
                for i in replaced_worker[name][tema]['отчёты']:
                    tema_report[tema][name]['отчёты'].append(i)


def worker_into_thems_2():
    for name in replaced_worker:
        for tema in replaced_worker[name]:
            if tema not in tema_report:
                tema_report[tema] = {}
                tema_report[tema][name] = replaced_worker[name][tema]
            elif name not in tema_report[tema]:
                tema_report[tema][name] = replaced_worker[name][tema]
            else:
                for i in replaced_worker[name][tema]:
                    tema_report[tema][name][i] += replaced_worker[name][tema][i]


# Сворачивает отчёты в компактный вид
def short_show_report(massiv):
    massiv_to_return = []
    for otchet in massiv:
        word = ''
        for b in otchet:
            if b.isdigit():
                word += b
            else:
                break
        massiv_to_return.append(word)

    return massiv_to_return


# Создает отчёт с именами
def make_excel():
    wb = opx.Workbook()
    ws_title = ['Выпущенные опытно-промышленные партии',
                'Изготовленные лабораторные образцы и компоненты',
                'Написанные отчеты', 'Выпущенные рецептуры и ТК', 'Проведенные промышленные нанесения']
    ws = wb.active
    ws_count_row = []
    for tema in tema_report:
        ws_title.insert(0, tema)
        ws.append(ws_title)
        ws_count_row.append(True)
        for name in tema_report[tema]:
            sum_obraz = len(tema_report[tema][name]['образцы']) + len(tema_report[tema][name]['плёнки'])
            sum_otchet = len(tema_report[tema][name]['отчёты'])
            current_list_to_append = [name]
            cell = ''
            for i in (short_show(tema_report[tema][name]['образцы']) +
                      short_show(tema_report[tema][name]['плёнки'])):
                cell += f'{i}, '
            if cell:
                cell = cell[:-2]
                cell += f'\n\nИтого: {sum_obraz}'
            current_list_to_append.append('')
            current_list_to_append.append(cell)
            cell = ''
            for i in short_show_report(tema_report[tema][name]['отчёты']):
                cell += f'{i}, '
            if cell:
                cell = cell[:-2]
                cell += f'\n\nИтого: {sum_otchet}'
            current_list_to_append.append(cell)
            for i in range(2):
                current_list_to_append.append(' ')
            ws.append(current_list_to_append)
            ws_count_row.append(True)
        ws_title.pop(0)
        for _ in range(2):
            ws.append([' '])
            ws_count_row.append(False)
        first_row = True
        for index, row, row_numb in zip(ws_count_row, ws, itertools.count(1, 1)):
            if index:
                for cell in row:
                    if first_row:
                        cell.font = opx.styles.Font(bold=True)
                        first_row = False
                    cell.border = thin_border
                    cell.alignment = Alignment(wrapText=True)

            else:
                ws.merge_cells(f'A{row_numb}:F{row_numb}')
                first_row = True

    ws.page_setup.paperSize = '9'

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 18
    wb.save('Otchet.xlsx')
    return os.getcwd() + '\Otchet.xlsx'

# Создает отчёт без имен
def make_excel_noname():
    ws_title = ['Темы', 'Выпущенные опытно-промышленные партии',
                'Изготовленные лабораторные образцы и компоненты',
                'Написанные отчеты', 'Выпущенные рецептуры и ТК', 'Проведенные промышленные нанесения']

    wb = opx.Workbook()
    ws = wb.active
    ws_count_row = []
    ws.append(ws_title)
    ws_count_row.append(True)
    for tema in tema_report:
        current_list_to_append = [tema]
        sum_obraz = 0
        sum_otchet = 0
        cell = ''
        cell2 = ''
        for name in tema_report[tema]:
            sum_obraz += len(tema_report[tema][name]['образцы']) + len(tema_report[tema][name]['плёнки'])
            sum_otchet += len(tema_report[tema][name]['отчёты'])
            for i in short_show(tema_report[tema][name]['образцы']):
                cell += f'{i}, '
            if cell:
                cell = cell[:-2] + '\n'
            for i in short_show(tema_report[tema][name]['плёнки']):
                cell += f'{i}, '
            for i in short_show_report(tema_report[tema][name]['отчёты']):
                cell2 += f'{i}, '

        if cell:
            cell = cell[:-2] + f'\n\nИтого: {sum_obraz}'
        if cell2:
            cell2 = cell2[:-2] + f'\n\nИтого: {sum_otchet}'

        current_list_to_append.append('')
        current_list_to_append.append(cell)
        current_list_to_append.append(cell2)
        for i in range(2):
            current_list_to_append.append('')
        ws.append(current_list_to_append)
        ws_count_row.append(True)

        for index, row, row_numb in zip(ws_count_row, ws, itertools.count(1, 1)):
            if index:
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(wrapText=True)
            else:
                ws.merge_cells(f'A{row_numb}:F{row_numb}')

    ws.page_setup.paperSize = '9'

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 18
    wb.save('Otchet_noname.xlsx')
    return os.getcwd() + '\Otchet_noname.xlsx'

# Заменяет темы правильными наименованиями
def replace_names_thems(theme_list, known_themes):
    global themes, unknown_thems, replaced_worker
    known_themes_local = []
    for tema in theme_list:
        for tema_2 in theme_list[tema]:
            known_themes_local.append(tema_2)
        if tema not in known_themes_local:
            known_themes_local.append(tema)
    unknown_thems = []
    replaced_worker = {}
    for name in worker:
        for tema_to_check in worker[name]:
            if tema_to_check in known_themes_local:
                for global_tema in theme_list:
                    if tema_to_check in theme_list[global_tema]:
                        if name in replaced_worker:
                            if global_tema in replaced_worker[name]:
                                for i in replaced_worker[name][global_tema]:
                                    replaced_worker[name][global_tema][i] += worker[name][tema_to_check][i]
                            else:
                                replaced_worker[name][global_tema] = worker[name][tema_to_check]
                        else:
                            replaced_worker[name] = {}
                            replaced_worker[name][global_tema] = worker[name][tema_to_check]

            else:
                if tema_to_check not in unknown_thems:
                    unknown_thems.append(tema_to_check)
                if name not in replaced_worker:
                    replaced_worker[name] = {}
                if tema_to_check not in replaced_worker[name]:
                    replaced_worker[name][tema_to_check] = {}
                replaced_worker[name][tema_to_check] = worker[name][tema_to_check]

    themes = known_themes_local + unknown_thems
    print(known_themes_local)
    print(unknown_thems)

    # for tema in themes.copy():
    #     for global_tema in theme_list:
    #         if tema in theme_list[global_tema]:
    #             themes.pop(themes.index(tema))


def save_data():
    data_to_save = {'svod_data': svod_data, 'report_data': report_data, 'production_data': production_data,
                    'report_row': report_row, 'svod_row': svod_row, 'production_row': production_row}
    with open('data.otchet', 'wb') as f:
        pickle.dump(data_to_save, f)
    # with open('data2.otchet', 'wb') as f:
    #     pickle.dump(production_data, f)


def load_data():
    global svod_data, production_data, report_data, report_row, svod_row, production_row
    if os.path.exists('data.otchet'):
        with open('data.otchet', 'rb') as f:
            data_to_load = pickle.load(f)

            svod_data = data_to_load['svod_data']
            svod_row = data_to_load['svod_row']
            production_data = data_to_load['production_data']
            production_row = data_to_load['production_row']
            report_data = data_to_load['report_data']
            report_row = data_to_load['report_row']





    for name in production_data:
        print(name, len(production_data[name]))

    pass



if __name__ == '__main__':
    for i in conductor('Сводная таблица.xlsm', 'Общий перечень продукции ОВНТ.xlsx', 'Общий перечень отчётов.xlsm'):
        pass


    worker_into_thems()
    make_excel()
